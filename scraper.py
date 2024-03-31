from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
from openpyxl import load_workbook


def scrape(url, doc):
    src1 = requests.get(url, headers={
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:120.0) Gecko/20100101 Firefox/120.0'})

    soup = BeautifulSoup(src1.content, 'lxml')
    cards = soup.find_all('div', class_="propertyCard-wrapper")

    for card in cards:
        address = card.find('address').text.strip()
        try:
            price1 = card.find('div', class_='propertyCard-rentalPrice-primary').text.strip()
        except AttributeError:
            price1 = card.find('div', class_='propertyCard-priceValue').text.strip()

        price1 = price1.split()
        price1 = price1[0]

        contact1 = card.find('p',
                             class_="propertyCard-contactsItemDetails propertyCard-contactsItemDetails--phone").text.replace(
            'Local call rate', '').replace(' ', '').strip()
        info = card.find('h2', class_="propertyCard-title").text.strip()
        target = card.find('a', class_="propertyCard-link")['href']
        page = f'https://www.rightmove.co.uk{target}'

        excel_file_path = f'{doc}.xlsx'
        try:
            wb = load_workbook(excel_file_path)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(['Contact', 'Address', 'Price', 'Info', 'Page'])
        ws.append([contact1, address, price1, info, page])
        wb.save(excel_file_path)


index = ''
print("""════════════════════════════════""")
url = input("URL to scrape: ")
if "&index=" in url:
    url = url.split("&index=")
    url = url[0]
    print(url)
print("""════════════════════════════════""")
doc = input("CSV file name to append/create: ")
if ".xlsx" in doc:
    doc = doc.replace(".xlsx", '')
print("""════════════════════════════════""")
maxp = int(input("Pages to search: "))
print("Scraping to CSV file, please do not cancel...")

scrape(url, doc)

i = 0
j = 1

while j <= maxp:
    print("""════════════════════════════════""")
    print("🔃 Scraping page: ", j, "/", maxp)

    i += 24
    j += 1
    index = f'&index={i}'
    url += index
    try:
        scrape(url, doc)

        print("✅ Scraped page: ", j - 1)

    except:
        print("❌ Process finished at page: ", j - 1)
        break
